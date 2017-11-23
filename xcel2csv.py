#!/usr/bin/env python

import sys
from operator import attrgetter
from openpyxl import load_workbook
from toolz.functoolz import compose

value = attrgetter('value')
def format_row(row, sep=', '):
    """Return row as 'sep' separated string."""
    format_item = compose(str, value)
    return sep.join(map(format_item, row))

def debug_row(row):
    """Some rows are corrupt - do best to print them out."""
    for i in row:
        print value(i)

def rows_in_workbook(file):
    """Return iterator over rows of the first sheet in ifile."""
    wb = load_workbook(ifile)
    return wb.active.iter_rows()

def copy_to_csv(ifile, fo): 
    """Copy spreadsheet ifile to file object fo as csv."""
    for row in rows_in_workbook(ifile):
        try:
            fo.write(format_row(row) + '\n')
        except UnicodeEncodeError:
            debug_row(row)
            break
   
ifile = sys.argv[1]
ofile = ifile + '.csv'

with open(ofile, 'w') as fo:
    copy_to_csv(ifile, fo)
    
