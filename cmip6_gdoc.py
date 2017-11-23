"""
Provides an interface on the xcel version of the google spreadsheet.
"""

from __future__ import print_function

import sys
import re
from operator import attrgetter
from collections import namedtuple
from functools import wraps
from string import maketrans

import openpyxl
MODEL = 'HadGEM3' # temporary hardcode - should be an option

def open(fname, filt=True):
    spsh = openpyxl.load_workbook(fname, read_only=True)
    
    return RequestWithMappings(spsh, filt)

STASH_RANGE_PATTERN = re.compile(r'm01s(\d\d)i(\d\d\d)\:i(\d\d\d)')
STASH_PATTERN = re.compile(r'm01s(\d\d)i(\d\d\d)')
STASH_FMT = "m01s{0:02d}i{1:03d}"

def stashlist_from_mapping_entry(entry):
    """
    Strip out a list of the stash codes involved in a mapping.
    """
    retval = []
	
    if entry is None:
        return None
	
    try:
        range_patterns = STASH_RANGE_PATTERN.findall(entry)
    except TypeError:
        warnings.warn("ignoring entry %s"%entry)
        return None
	
    if range_patterns is not None:
        for j in range_patterns:
            section, item1, item2 = [int(i) for i in j]
            retval += [STASH_FMT.format(section, i) for i in range(item1, item2 + 1)]
	
    single_patterns = STASH_PATTERN.findall(entry)
    if single_patterns is not None:
        for j in single_patterns:
            section, item = [int(i) for i in j]
            retval.append(STASH_FMT.format(section, item))
	   
    # remove duplicates and sort
    retval = sorted(list(set(retval)))
    return retval

def _to_attr(value):
    """
    Return a value as in a form suitable for use as attribute.
    
    >>> _to_attr('Variable mapping')
    'variable_mapping'

    >>> _to_attr("Notes (this doesn't)")
    'notes_this_doesnt'

    >>> _to_attr("replace:colons")
    'replace_colons'
    """
    table = maketrans(' :', '__')
    return value.lower().translate(table, '()\"\'')

_TITLES = ['cmor_label', 'miptable', 
           'cell_methods', 'dimension',
           'units', 'realm', 'priority',
           'frequency', 'UKESM_component',
           'Variable_mapping', 'Plan',
           "Notes (this doesn't go in the metadata)",
           "Manual edit",
           "requestVarGroup membership (lists of mip:rvg label)",
           "cf_std_name",
           "requesting_mips",
           "Comment (this goes into file metadata)",
           "Ticket", 
           "last_update", "title", "positive"]

_FIELDS =  map(_to_attr, _TITLES)

class Request(namedtuple('Request', _FIELDS)):
    """
    Object representing a MIP requested variable.

    Infers stash_codes_needed from the mapping.

    Examples
    --------
    
    To help set up test data for examples we need a small function:
    >>> def new_sample(): 
    ...     result =  [None for i in range(18)]
    ...     result[3] = 'longitude'
    ...     return result

    >>> sample = new_sample()

    An empty mapping returns None
    >>> a = Request(*sample)
    >>> a.stash_codes_needed == None
    True
    >>> a.stash_codes
    []

    >>> sample[0] = 'var'
    >>> sample[1] = 'table'
    >>> sample[9] = 'm01s01i001'
    >>> sample[6] = ' 2'
    >>> a = Request(*sample)
    >>> a.mip_id
    'table_var'
    >>> a.stash_codes_needed
    'm01s01i001'
    >>> a.stash_codes
    ['m01s01i001']
    >>> a.priort
    2

    >>> sample[9] = 'm01s01i001/m01s01i002'
    >>> a = Request(*sample)
    >>> a.stash_codes_needed
    'm01s01i001,m01s01i002'
    >>> a.stash_codes
    ['m01s01i001', 'm01s01i002']

    >>> sample = new_sample()
    >>> sample[9] = 'm01s08i248 * m01s03i395 * 100'
    >>> a = Request(*sample)
    >>> a.stash_codes
    ['m01s03i395', 'm01s08i248']

    Mappings without stash should return None
    >>> sample[9] = 'DMS_SURF * 1e-6'
    >>> a = Request(*sample)
    >>> a.stash_codes_needed == None
    True
    >>> a.stash_codes
    []

    A mapping with a HadGEM3 specific mapping should return it
    >>> sample[9] = 'm01s19i001'
    >>> sample[11] = 'HadGEM3_variable_mapping:veg(m01s03i317,m01s00i505,vegClass="bareSoil"):notes:'
    >>> a = Request(*sample)
    >>> a.stash_codes
    ['m01s00i505', 'm01s03i317']

    Some mappings are not for HadGEM3.  These include:
       1. section 19 diagnostics
    >>> sample[9] = 'm01s19i001'
    >>> sample[11] = None
    >>> a = Request(*sample)
    >>> a.is_hadgem3
    False

       2. ukca or ogbc diagnostics
    >>> sample[9] = 'm01s50i100'
    >>> sample[8] = 'chemistry'
    >>> a = Request(*sample)
    >>> a.is_hadgem3
    False

    >>> sample = new_sample()
    >>> sample[8] = 'obgc'
    >>> a = Request(*sample)
    >>> a.is_hadgem3
    False

      3. multi dimensional CO2 or surface CO2 emissions
    >>> sample = new_sample()
    >>> sample[9] = 'm01s00i251'
    >>> a = Request(*sample)
    >>> a.is_hadgem3
    False

    >>> sample = new_sample()
    >>> sample[9] = 'm01s00i252'
    >>> a = Request(*sample)
    >>> a.is_hadgem3
    False

    If there is a 'manual edit' to the spread sheet then a 'plan' of 'request-error' should be
    set to 'available'.

    The usual case is with an empty 'manual edit' column
    >>> sample[8] = ''
    >>> sample[10] = 'request-error'
    >>> sample[12] = ''
    >>> a = Request(*sample)
    >>> a.inferred_plan
    'request-error'

    But occasionally we see an entry in 'manual edit':
    >>> sample[12] = 'cell_methods corrected'
    >>> a = Request(*sample)
    >>> a.inferred_plan
    'available'

    If the request asks for a STASH diagnostic on model levels then the orography will
    need to be present.  To make it simple for subsequent processing add orography as
    an extra stash code.
    
    >>> sample = new_sample()
    >>> sample[9] = 'm01s01i001'
    >>> sample[3] = 'alev'
    >>> a = Request(*sample)
    >>> a.stash_codes
    ['m01s01i001', 'm01s00i033']

    The request group is in a column with a long name - so we support an abbreviation
    >>> sample = new_sample()
    >>> sample[13] = '"CFMIP:CFday-3d", "RFMIP:RFMIP-AeroIrf"'
    >>> a = Request(*sample)
    >>> a.groups
    ['CFMIP:CFday-3d', 'RFMIP:RFMIP-AeroIrf']

    If no request groups are found then return list with an empty string
    >>> sample = new_sample()
    >>> sample[13] = ''
    >>> a = Request(*sample)
    >>> a.groups
    ['']

    Priority can be overriden by a note of the form MO_priority
    >>> sample = new_sample()
    >>> sample[11] = 'MO_priority:1:'
    >>> sample[6] = ' 2'
    >>> a = Request(*sample)
    >>> a.priort
    1

    """
    _MODEL_LEVELS = 'alev' # (this may need extending)
    _OROG_STASH = 'm01s00i033'
    _HADGEM3 = re.compile('HadGEM3_variable_mapping:(.*?):')
    _MO_PRIORITY = re.compile('MO_priority:(.*?):')

    @property
    def mip_id(self):
        return self.miptable + '_' + self.cmor_label

    @property
    def _priority_override(self):
        result = False
        if self._notes:
            result = self._MO_PRIORITY.search(self._notes)
        return result

    @property
    def priort(self):
        priority = self.priority
        if self._priority_override:
            priority = self._priority_override.group(1)
        return int(priority)

    @property
    def is_hadgem3(self):
        has_ukesm1_only_component = self.ukesm_component in ('chemistry', 'obgc')
        has_section19 = len([x for x in self.stash_codes if 's19' in x]) > 0 
        has_co2_ukesm1 = len([x for x in self.stash_codes 
                                 if x in ('m01s00i251', 'm01s00i252')]) > 0

        return not any((has_ukesm1_only_component,
                        has_section19,
                        has_co2_ukesm1))
            

    @property
    def stash_codes_needed(self):
        result = None
        if self.stash_codes:
            result = ','.join(self.stash_codes)
        return result

    @property
    def stash_codes(self):
        result = []
        if self._variable_mapping:
            result = stashlist_from_mapping_entry(self._variable_mapping)
            if self._MODEL_LEVELS in self.dimension:
                result.append(self._OROG_STASH)
        return result

    @property
    def inferred_plan(self):
        return self.plan if not self.manual_edit else 'available'

    @property
    def groups(self):
        grps = self.requestvargroup_membership_lists_of_mip_rvg_label
        return str(grps).translate(None, '" ').split(',')  # sort out formatting

    @property
    def _notes(self): # add in a short hand
        return self.notes_this_doesnt_go_in_the_metadata
 
    @property
    def _hadgem3_in_notes(self):
        result = False
        if self._notes:
            result = self._HADGEM3.search(self._notes)
        return result

    @property
    def _variable_mapping(self):
        result = self.variable_mapping
        if MODEL == 'HadGEM3' and self._hadgem3_in_notes:
            result = self._hadgem3_in_notes.group(1)
        return result

def _strip(val):
    """
    Strip surrounding whitespace if possible, otherwise return argument.

    >>> _strip(' this thing ')
    'this thing'

    >>> _strip(10)
    10
    """
    try:
        result = val.strip()
    except AttributeError:
        result = val
    return result

class RecordReader(object):
    def __init__(self, entries):
        self._entries = entries
        self._find_indexes()

    def __call__(self, row):
        vals = (row[index].value for index in self._indices)
        args = map(_strip, vals)
        return Request(*args)
        
    def _find_indexes(self):
        self._indices = []
        for title in _TITLES:
            self._indices.append(self._entries.index(title))

def log_filter(func):
    """
    Decorator to log that a record is filtered out by a filter.

    Example
    -------
    >>> from collections import namedtuple
    >>> myrecord = namedtuple('myrecord', ['miptable', 'cmor_label'])
    >>> dummy = myrecord('Amon', 'tas')

    >>> def yes(record): return True
    >>> wyes = log_filter(yes)
    >>> wyes(dummy)
    True
    >>> def no(record): return False
    >>> wno = log_filter(no)
    >>> wno(dummy)
    FILTER "no": Amon tas
    False
    """
    @wraps(func)
    def _logged(record):
        result = func(record)
        if not result:
            fmt = 'FILTER "{}": {r.miptable} {r.cmor_label}'
            print(fmt.format(func.__name__, r=record))
        return result
    return _logged

def available(record):
    """
    Return True if the record is availiable.

    Examples
    --------
    >>> from collections import namedtuple
    >>> myrecord = namedtuple('myrecord', ['inferred_plan'])
    >>> available(myrecord(inferred_plan=None))
    False

    >>> available(myrecord(inferred_plan=' do-not-produce '))
    False

    >>> available(myrecord(inferred_plan='available'))
    True
    """
    if record.inferred_plan != None:
        return record.inferred_plan.strip() in ('available', 'post-process',
                                                'vn10.6',
                                                'vn10.6.1',
                                                'vn10.7')
    else:
        return False

@log_filter
def not_rogue(record):
    """
    Return True if diagnostic is not rogue.

    'Rogue' means 'somethign we can't process for some
    hopefully-ephemeral reason: this filter should eventually not do
    anything.
    """
    return True

@log_filter
def good_freq(record):
    freqs = ('fx')
    return record.frequency not in freqs

def not_site(record):
    """
    Returns True if site is not in the dimension.

    Examples
    --------
    >>> from collections import namedtuple
    >>> myrecord = namedtuple('myrecord', ['dimension'])
    >>> not_site(myrecord(dimension='time latitude'))
    True

    >>> not_site(myrecord(dimension='time site'))
    False
    """
    return 'site' not in record.dimension

@log_filter
def hadgem3(record):
    """
    Returns True if valid for hadgem3.
    """
    return record.is_hadgem3

def required_components(record):
    required = ('aerosol',
                'atmos-physics',
                'boundary layer',
                'carbon',
                'cftables',
                'chemistry',
                'cloud',
                'dust',
                'icesheet',
                'land',
                'land-use',
                'obgc',
                'ocean',
                'radiation',
                'seaice',
                'snow-permafrost')
    return record.ukesm_component in required


def has_stash(record):
    """
    Return True if the record has a stash based mapping
    
    Examples
    --------
    >>> from collections import namedtuple
    >>> myrecord = namedtuple('myrecord', ['stash_codes_needed'])
    >>> has_stash(myrecord(stash_codes_needed='m01s01i001'))
    True

    >>> has_stash(myrecord(stash_codes_needed=None))
    False
    """
    return record.stash_codes_needed != None

def compound_filter(*args):
    """
    Returns a function that is a compound record filter.

    Examples
    --------

    >>> filter_func = compound_filter(lambda x: x, lambda x: x)
    >>> filter_func(True)
    True
    >>> filter_func(False)
    False

    >>> filter_func = compound_filter(lambda x: not x, lambda x: x)
    >>> filter_func(True)
    False
    """
    def filter(record):
        return all(afilter(record) for afilter in args)
    return filter

include = compound_filter(available,
                          good_freq,
                          not_site,
                          has_stash,
                          hadgem3,
                          not_rogue)

def still_valid(row):
    """
    Returns True if the row looks like it is valid.
    """
    result = True
    return True  # FIXME
    if not row[0].value: # removes empty rows - could be better
        result = False
    else:
        result = row[26].value == None or 'DELETE' not in row[26].value
    return result
 
class RequestWithMappings(object):
    def __init__(self, workbook, filt=True):
        self._sheet = workbook.get_sheet_by_name('Diagnostics')
        self._rows = self._sheet.rows
        self.titles = self._titles()
        self._reader = RecordReader(self.titles)
        records = (self._reader(row) for row in self._rows if still_valid(row))
        if filt:
            self.records = filter(include, records)
        else:
            self.records = list(records)

    def __iter__(self):
        return iter(self.records)

    def _titles(self):
        result = map(attrgetter('value'), next(self._rows))
        if not result[0]:  # for some reason the reading returns None?
            result[0] = 'cmor_label'
        return result

    def _records_for_table(self, table):
        return [rec for rec in self.records if rec.miptable == table]

    def tables(self):
        return [Table(table, self._records_for_table(table)) for table in self.table_names()]

    def table_names(self):
        return tuple(sorted(set(rec.miptable for rec in self.records)))
        
    # this is maintained for consistency - not sure it is really necessary
    def first_table_dim(self):
        return (self._sheet.max_row, self._sheet.max_column)
    

class Table(object):
    """
    A mip table read from the CMIP6 spread sheet
    """
    mo_priority_col = None # consistency with old interface

    def __init__(self, table_name, records):
        self.title = table_name
        self.records = records

    def cell_methods(self):
        return [rec.cell_methods for rec in self.records]

    def dimensions(self):
        return [rec.dimension for rec in self.records]

    def uniques(self):
        # a better method?
        return ['{}_{}'.format(rec.miptable, rec.cmor_label) for rec in self.records]
    
    def realms(self):
        return [rec.realm for rec in self.records]

    def times(self):
        return [rec.frequency for rec in self.records]
               
    def stashs(self):
        # would be better to infer from mapping
        return [rec.stash_codes_needed for rec in self.records]

    def cmors(self):
        return [rec.cmor_label for rec in self.records]

    def varnames(self):
        return self.cmors()
              
    def units(self):
        return [rec.units for rec in self.records]

    def cmip6_priorities(self):
        return [rec.priority for rec in self.records]

    def mo_priorities(self):
        raise NotImplementedError()

    def ukesm_components(self):
        return [rec.ukesm_component for rec in self.records]
